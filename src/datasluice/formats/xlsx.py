"""XLSX format reader (requires the ``openpyxl`` optional dependency)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from datasluice.exceptions import FormatError
from datasluice.formats.base import BaseFormatReader


class XLSXReader(BaseFormatReader):
    """Read XLSX spreadsheets into a list of dictionaries.

    Requires ``openpyxl``: install with ``pip install datasluice[xlsx]``.
    """

    format_name = "XLSX"

    def read(self, source: str | Path | bytes) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise FormatError("XLSX support requires 'openpyxl'. Install with: pip install datasluice[xlsx]") from exc

        if isinstance(source, bytes):
            import io

            wb = load_workbook(io.BytesIO(source), read_only=True, data_only=True)
        else:
            wb = load_workbook(source, read_only=True, data_only=True)

        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows)]
        except StopIteration:
            return []

        records: list[dict[str, Any]] = []
        for row in rows:
            records.append(dict(zip(headers, row, strict=False)))
        return records
